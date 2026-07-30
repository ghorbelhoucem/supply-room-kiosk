from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.auth import require_manager
from app.database import get_db
from app.models import Checkout, InventoryItem, Movement, User
from app.services.inventory import availability_label, available_qty

router = APIRouter(tags=["reports"])


@router.get("/reports/summary")
def report_summary(db: Session = Depends(get_db), _user: User = Depends(require_manager)):
    items = db.execute(select(InventoryItem).order_by(InventoryItem.name)).scalars().all()
    open_tx = (
        db.execute(select(Checkout).where(Checkout.returned_at.is_(None))).scalars().all()
    )
    overdue = []
    now = datetime.now(timezone.utc)
    for c in open_tx:
        if c.expected_return and c.expected_return < now:
            item = db.get(InventoryItem, c.item_id)
            overdue.append(
                {
                    "txId": c.tx_id,
                    "item": item.name if item else "?",
                    "personRole": c.person_role,
                    "expectedReturn": c.expected_return.isoformat(),
                }
            )
    low = [
        {
            "item": it.name,
            "quantity": it.qty_on_hand,
            "available": available_qty(db, it),
            "reorder_min": it.reorder_min,
        }
        for it in items
        if available_qty(db, it) <= it.reorder_min
    ]
    return {
        "ok": True,
        "open_count": len(open_tx),
        "overdue": overdue,
        "low_stock": low,
    }


@router.get("/exports/inventory.xlsx")
def export_inventory_xlsx(db: Session = Depends(get_db), _user: User = Depends(require_manager)):
    wb = Workbook()

    ws = wb.active
    ws.title = "Inventory"
    ws.append(["SKU", "Item", "Category", "Qty On Hand", "Available", "Reorder Min", "Below Min", "Barcode"])
    items = db.execute(select(InventoryItem).order_by(InventoryItem.name)).scalars().all()
    for it in items:
        avail = available_qty(db, it)
        ws.append(
            [
                it.sku,
                it.name,
                it.category.value,
                it.qty_on_hand,
                avail,
                it.reorder_min,
                "YES" if avail <= it.reorder_min else "NO",
                it.barcode or "",
            ]
        )

    ws2 = wb.create_sheet("Open Checkouts")
    ws2.append(["TxId", "Item", "Taken By", "Taken At", "Expected Return"])
    opens = db.execute(select(Checkout).where(Checkout.returned_at.is_(None))).scalars().all()
    for c in opens:
        item = db.get(InventoryItem, c.item_id)
        ws2.append(
            [
                c.tx_id,
                item.name if item else "",
                c.person_role,
                c.taken_at.isoformat() if c.taken_at else "",
                c.expected_return.isoformat() if c.expected_return else "",
            ]
        )

    ws3 = wb.create_sheet("Recent Movements")
    ws3.append(["When", "Type", "Item", "Qty", "Actor", "Reason", "TxId"])
    moves = (
        db.execute(select(Movement).order_by(Movement.created_at.desc()).limit(500)).scalars().all()
    )
    for m in moves:
        ws3.append(
            [
                m.created_at.isoformat() if m.created_at else "",
                m.movement_type.value,
                m.item_name,
                m.qty,
                m.actor,
                m.reason or "",
                m.related_tx_id or "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"inventory-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
